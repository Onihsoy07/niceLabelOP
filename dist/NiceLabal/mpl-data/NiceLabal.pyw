import sys
import os
import openpyxl as op
import pandas as pd
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import uic

class ExcelTrans:
    def __init__(self):
        pass

# Excel 제품명 쪼개기
    def itemSepa(self, item):
        name = ''
        ratting = ''
        acc = ''
        note = ''
        attFlag = 0
        if 'ATT' in item:
            attFlag = 1
            item = item.replace('ATT', '')
        item = item.translate(item.maketrans({'[' : '', ']' : ''}))
        itemList = item.split()
        if attFlag == 1:
            itemList[0] = itemList[0] + ' (ATT)'
        if len(itemList) == 2:
            name = itemList[0]
            ratting = itemList[1]
        elif len(itemList) == 3:
            if itemList[-1] == 'E':
                name = itemList[0]
                ratting = itemList[1]
                note = itemList[2]
            elif len(itemList[-1]) == 1:
                name = itemList[0]
                ratting = itemList[1]
                acc = itemList[2]
            elif '+' in itemList[-1]:
                name = itemList[0]
                ratting = itemList[1]
                acc = itemList[2]
            else:
                name = itemList[0]
                ratting = itemList[1]
                note = itemList[2]
        elif len(itemList) == 4:
            name = itemList[0]
            ratting = itemList[1]
            acc = itemList[2]
            note = itemList[3]
        return name, ratting, acc, note

#Excel 제품명 정리 저장
    def setItem(self, path):
        itemCol = 2
        nameCol = 3
        rattingCol = 4
        accCol = 5
        noteCol = 6
        excelWB = op.load_workbook(path)
        excelWS = excelWB.active
        for _ in range(4):  excelWS.insert_cols(itemCol + 1)
        excelWS.cell(row=1, column=nameCol).value = 'name'
        excelWS.cell(row=1, column=rattingCol).value = 'ratting'
        excelWS.cell(row=1, column=accCol).value = 'acc'
        excelWS.cell(row=1, column=noteCol).value = 'note'
        maxRow = excelWS.max_row
        maxColumn = excelWS.max_column
        for i in range(2, maxRow + 1):
            name, ratting, acc, note = self.itemSepa(excelWS.cell(row=i, column=itemCol).value)
            excelWS.cell(row=i, column=nameCol).value = name
            excelWS.cell(row=i, column=rattingCol).value = ratting
            excelWS.cell(row=i, column=accCol).value = acc
            excelWS.cell(row=i, column=noteCol).value = note
        excelWB.save(path)
        excelWB.close()

#Excel 인쇄수량 데이터 변환
    def setPrint(self, path):
        dataCols = [3, 4, 13, 30]
        startRow = 3
        quantityColumn = 3
        packagingColumn = 4
        printQuantity = 5

        excelAddress = path
        newExcel = excelAddress[0:-5] + 'NiceLabel용.xlsx'
        wb = op.load_workbook(excelAddress)
        ws = wb.active
        maxRow = ws.max_row
        maxColumn = ws.max_column

        for i in range(maxColumn, 0, -1):
            if not i in dataCols:
                ws.delete_cols(i)

        ws.cell(column=printQuantity, row=1).value = '인쇄수량'
        maxColumn = ws.max_column
        for i in range(maxRow, (startRow - 1), -1):
            if (ws.cell(row=i, column=quantityColumn).value % ws.cell(row=i, column=packagingColumn).value) != 0:
                if (ws.cell(row=i, column=quantityColumn).value // ws.cell(row=i, column=packagingColumn).value) == 0:
                    ws.cell(row=i, column=packagingColumn).value = ws.cell(row=i, column=quantityColumn).value
                    ws.cell(row=i, column=printQuantity).value = 1
                else:
                    ws.insert_rows(i + 1)
                    for j in range(1, maxColumn):
                        ws.cell(row=(i + 1), column=j).value = ws.cell(row=i, column=j).value
                    ws.cell(row=i, column=printQuantity).value = ws.cell(row=i, column=quantityColumn).value // ws.cell(
                        row=i, column=packagingColumn).value
                    ws.cell(row=i + 1, column=packagingColumn).value = ws.cell(row=i,
                                                                               column=quantityColumn).value % ws.cell(
                        row=i, column=packagingColumn).value
                    ws.cell(row=i + 1, column=printQuantity).value = 1
            else:
                ws.cell(row=i, column=printQuantity).value = ws.cell(row=i, column=quantityColumn).value // ws.cell(
                    row=i, column=packagingColumn).value
        ws.delete_rows(2)
        wb.save(newExcel)
        wb.close()
        return newExcel

    def transXlsx(self, path):
        excelPath = path
        if path[-3:] == 'xls':
            excelPath = path[:-3] + 'xlsx'
            excelDF = pd.read_excel(path)
            excelDF.to_excel(excelPath, index=False)
        return excelPath


class ColorSetting(QDialog):
    def __init__(self):
        super().__init__()

        self.colorTxt = 'Color1'
        self.num = 3

        self.ui = uic.loadUi("ColorSetting.ui", self)
        self.setWindowTitle('ColorSetting')
        self.colorComboAction('Color1')

        self.btn_itemAdd.clicked.connect(self.itemAdd)
        self.btn_itemRemove.clicked.connect(self.itemRemove)
        self.btn_colorItemSave.clicked.connect(self.colorItemSave)

        for i in range(1, self.num + 1):
            comboBoxName = 'color' + str(i)
            self.colorComboBox.addItem(comboBoxName)

        self.colorComboBox.activated[str].connect(self.colorComboAction)

#ColorSetting Window 시작
        self.show()

    def listWdgetData(self):
        row = 0
        data = []
        while self.itemList.item(row):
            if self.itemList.item(row):
                data.append(self.itemList.item(row).text())
                row += 1
        return row, data

    def colorItemSave(self):
        row, data = self.listWdgetData()
        colorItemList = ''
        with open(self.colorTxt + '.txt', 'w') as colorTxt:
            for item in data:
                if not colorItemList:
                    colorItemList += item
                else:
                    colorItemList = colorItemList + ' ' + item
            colorTxt.write(colorItemList)

    def colorComboAction(self, text):
        self.itemList.clear()
        self.colorTxt = text
        with open(text + '.txt', 'r') as colorTxt:
            dataList = colorTxt.read().split()
            for dataTxt in dataList:
                self.itemList.addItem(dataTxt.strip())

    def itemAdd(self):
        self.itemtext = self.itemText.text()
        self.itemList.addItem(self.itemtext)
        self.num += 1
        print(self.num)
    def itemRemove(self):
        removeItemRow = self.itemList.currentRow()
        self.itemList.takeItem(removeItemRow)


class MyWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.excelTrans = ExcelTrans()
        self.windowWeight = 600
        self.windowHight = 320
        self.setWindowTitle("NiceLabel Excel 변환")
        self.f = open('NiceLabelLocate.txt', 'r')
        self.appPath = self.f.read()
        self.excelFileFlag = ''
        self.setupUI()
        self.f.close()

    def setupUI(self):
        self.setGeometry(500, 250, self.windowWeight, self.windowHight)
        self.setFixedSize(self.windowWeight, self.windowHight)

        self.filelabel = QLabel('', self)
        self.filelabel.move(135, 30)
        self.filelabel.resize(500, 30)

        self.nclabel = QLabel(self.appPath, self)
        self.nclabel.move(135, 120)
        self.nclabel.resize(500, 30)

        self.excelTransMassage = QLabel('', self)
        self.excelTransMassage.move(135, 60)
        self.excelTransMassage.resize(500, 30)

        self.appOpenMassage = QLabel('', self)
        self.appOpenMassage.move(135, 150)
        self.appOpenMassage.resize(500, 30)

#메뉴바 만들기
        setMenu = QAction('set', self)
        setMenu.triggered.connect(self.setAction)

        menubar = self.menuBar()
        colormenu = menubar.addMenu('&Color')
        colormenu.addAction(setMenu)

#메인메뉴 버튼
        btn_openfile = QPushButton("Excel File", self)
        btn_openfile.move(30, 30)
        btn_openfile.clicked.connect(self.btn_openFile)

        btn_exceltrans = QPushButton("변환", self)
        btn_exceltrans.move(30, 60)
        btn_exceltrans.clicked.connect(self.btn_excelTrans)

        btn_exit = QPushButton("종료", self)
        btn_exit.move(30, 210)
        btn_exit.clicked.connect(QCoreApplication.instance().quit)

        btn_apppath = QPushButton("NiceLabal File", self)
        btn_apppath.move(30, 120)
        btn_apppath.clicked.connect(self.btn_appPath)

        btn_openapp = QPushButton("Open", self)
        btn_openapp.move(80, 150)
        btn_openapp.resize(50, 30)
        btn_openapp.clicked.connect(self.btn_openApp)

        btn_apppathsave = QPushButton("Save", self)
        btn_apppathsave.move(30, 150)
        btn_apppathsave.resize(50, 30)
        btn_apppathsave.clicked.connect(self.btn_appPathSave)

#메뉴바 이벤트 테스트
    def setAction(self):
        colorSetting = ColorSetting()
        colorSetting.exec_()

#exe파일 위치 저장
    def btn_appPathSave(self):
        with open('NiceLabelLocate.txt', 'w') as f:
            if self.appPath:
                f.truncate()
                f.write(self.appPath)
                self.appOpenMassage.setText('저장이 완료되었습니다.')
            else:
                self.appOpenMassage.setText('저장할 파일이 없습니다.')

#exe파일 열기
    def btn_openApp(self):
        try:
            if self.appPath:
                os.popen(self.appPath)
                self.appOpenMassage.setText('파일이 열립니다.')
            else:
                self.appOpenMassage.setText('열 파일이 없습니다.')
        except:
            self.appOpenMassage.setText('파일이 이상합니다.')

#exe 파일 위치 불러오기
    def btn_appPath(self):
        self.nicefile = QFileDialog.getOpenFileName(self, 'NiceLabel File', './')
        self.appPath = self.nicefile[0]
        self.nclabel.setText(self.appPath)

#Excel 파일 위치 불러오기
    def btn_openFile(self):
        self.openfile = QFileDialog.getOpenFileName(self, 'Open file', './')
        self.excelFileFlag = self.openfile[0]
        self.filelabel.setText(self.excelFileFlag)

#Excel 데이터 변환
    def btn_excelTrans(self):
        if self.excelFileFlag:
            excelPath = self.excelTrans.transXlsx(self.openfile[0])
            excelPath = self.excelTrans.setPrint(excelPath)
            self.excelTrans.setItem(excelPath)
            self.excelTransMassage.setText('변환이 완료되었습니다.')
        else:
            self.excelTransMassage.setText('파일이 없습니다.')


if __name__ == "__main__":
    app = QApplication(sys.argv)
    mywindow = MyWindow()
    mywindow.show()
    app.exec_()
