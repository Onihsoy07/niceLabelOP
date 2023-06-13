
import logging
import sys
import os
import openpyxl as op
import pandas as pd
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import uic


def makeLogger(name=None):
    # 1 logger instance를 만든다.
    logger = logging.getLogger(name)

    if len(logger.handlers) > 0:
        return logger

    # 2 logger의 level을 가장 낮은 수준인 DEBUG로 설정해둔다.
    logger.setLevel(logging.DEBUG)

    # 3 formatter 지정
    formatter = logging.Formatter("[%(asctime)-10s] (줄 번호: %(lineno)d) %(name)s:%(levelname)s - %(message)s")

    # 4 handler instance 생성
    console = logging.StreamHandler()
    file_handler = logging.FileHandler(filename="INFO.log")

    # 5 handler 별로 다른 level 설정
    console.setLevel(logging.INFO)
    file_handler.setLevel(logging.DEBUG)

    # 6 handler 출력 format 지정
    console.setFormatter(formatter)
    file_handler.setFormatter(formatter)

    # 7 logger에 handler 추가
    logger.addHandler(console)
    logger.addHandler(file_handler)

    return logger


class ExcelTrans:
    def __init__(self):
        self.logger = makeLogger()

        self.kaSetChangeDate = ''
        self.domesticKaItemRowDict = {}
        self.overseasKaItemRowDict = {}

    def kaItemRowDictSet(self):
        try:
            kaPreXlsx = op.load_workbook('KASETPRE.xlsx')
            domesticPreSheet = kaPreXlsx['국내 데이터 전처리']
            overseasPreSheet = kaPreXlsx['수출 데이터 전처리']
            dItem = domesticPreSheet.cell(row=1, column=1).value.split()
            dRow = domesticPreSheet.cell(row=2, column=1).value.split()
            oItem = overseasPreSheet.cell(row=1, column=1).value.split()
            oRow = overseasPreSheet.cell(row=2, column=1).value.split()

            for i in range(len(dItem)):
                rowList = [int(dRow[i]), int(dRow[i+1])]
                self.domesticKaItemRowDict[dItem[i]] = rowList

            for i in range(len(oItem)):
                rowList = [int(oRow[i]), int(oRow[i+1])]
                self.overseasKaItemRowDict[dItem[i]] = rowList

            print(self.domesticKaItemRowDict)
            print(self.overseasKaItemRowDict)

            kaPreXlsx.close()

        except Exception as e:
            print(e)

    def itemKaDomesticSet(self, item):
        maxWeight = {}
        kaPreXlsx = op.load_workbook('KASET.xlsx')
        domesticPreSheet = kaPreXlsx['국내 데이터 전처리']
        itemRow = domesticPreSheet[item[:3]]
        for i in range(itemRow[0], itemRow[1]):
            if domesticPreSheet.cell(row=i, column=1).value in item:
                if domesticPreSheet.cell(row=i, column=2).value and (domesticPreSheet.cell(row=i, column=2).value in item):
                    



    def itemKaOverseasSet(self, item):
        kaPreXlsx = op.load_workbook('KASET.xlsx')
        overseasPreSheet = kaPreXlsx['수출 데이터 전처리']



#KA 분류 저장                                                수정필요####################
    def itemKaSave(self, path):
        if self.kaSetChangeDate != os.path.getmtime('KASET.xlsx'):
            self.kaDataPreSet()
        wb = op.load_workbook(path)
        ws = wb.active
        itemColuom = 2
        kaColuom = 111111
        maxRow = ws.max_row
        for i in range(2, maxRow):
            #ws.cell(row=i, column=3).value = et.itemKa(ws.cell(row=i, column=2).value)
            pass
        wb.save(path)
        wb.close()

    #KA 데이터 전처리
    def kaDataPreSet(self):
        kaSetXlsx = op.load_workbook('KASET.xlsx')
        kaSetXlsx.create_sheet('국내 데이터 전처리', 2)
        kaSetXlsx.create_sheet('수출 데이터 전처리', 3)
        domesticSheet = kaSetXlsx['국내']
        domesticPreSheet = kaSetXlsx['국내 데이터 전처리']
        overseasSheet = kaSetXlsx['수출']
        overseasPreSheet = kaSetXlsx['수출 데이터 전처리']
        dFirstNameList = []
        dRowList = []
        oFirstNameList = []
        oRowList = []
        dpcRow = 3
        opcRow = 3

        #국내 KA 데이터 전처리
        while True:
            domesticMaxRow = domesticSheet.max_row
            if domesticMaxRow <= 2:
                dRowList.append(domesticPreSheet.max_row + 1)
                domesticPreSheet.cell(row=1, column=1).value = ' '.join(dFirstNameList)
                domesticPreSheet.cell(row=2, column=1).value = ' '.join(map(str, dRowList))
                break
            firstName = domesticSheet.cell(row=3, column=2).value[:3]
            dFirstNameList.append(firstName)
            dRowList.append(dpcRow)
            for i in range(domesticMaxRow, 2, -1):
                if firstName in domesticSheet.cell(row=i, column=2).value:
                    domesticPreSheet.cell(row=dpcRow, column=1).value = domesticSheet.cell(row=i, column=2).value[3:]
                    domesticPreSheet.cell(row=dpcRow, column=2).value = domesticSheet.cell(row=i, column=3).value
                    domesticPreSheet.cell(row=dpcRow, column=3).value = domesticSheet.cell(row=i, column=4).value
                    domesticSheet.delete_rows(i, 1)
                    dpcRow = dpcRow + 1

        #수출 KA 데이터 전처리
        while True:
            overseasMaxRow = overseasSheet.max_row
            if overseasMaxRow <= 2:
                oRowList.append(overseasPreSheet.max_row + 1)
                overseasPreSheet.cell(row=1, column=1).value = ' '.join(oFirstNameList)
                overseasPreSheet.cell(row=2, column=1).value = ' '.join(map(str, oRowList))
                break
            firstName = overseasSheet.cell(row=3, column=2).value[:3]
            oFirstNameList.append(firstName)
            oRowList.append(opcRow)
            for i in range(overseasMaxRow, 2, -1):
                if firstName in overseasSheet.cell(row=i, column=2).value:
                    overseasPreSheet.cell(row=opcRow, column=1).value = overseasSheet.cell(row=i, column=2).value[3:]
                    overseasPreSheet.cell(row=opcRow, column=2).value = overseasSheet.cell(row=i, column=3).value
                    overseasPreSheet.cell(row=opcRow, column=3).value = overseasSheet.cell(row=i, column=4).value
                    overseasSheet.delete_rows(i, 1)
                    opcRow = opcRow + 1

        kaSetXlsx.save('KASETPRE.xlsx')
        kaSetXlsx.close()

    # Excel 제품명 쪼개기
    def itemSepa(self, item):
        name = ''
        ratting = ''
        acc = ''
        note = ''
        attFlag = 0
        if 'ATT' in item:
            attFlag = 1
            item = item.replace('ATT', '')
        item = item.translate(item.maketrans({'[' : '', ']' : ''}))
        itemList = item.split()
        if attFlag == 1:
            itemList[0] = itemList[0] + ' (ATT)'
        if len(itemList) == 2:
            name = itemList[0]
            ratting = itemList[1]
        elif len(itemList) == 3:
            if itemList[-1] == 'E':
                name = itemList[0]
                ratting = itemList[1]
                note = itemList[2]
            elif len(itemList[-1]) == 1:
                name = itemList[0]
                ratting = itemList[1]
                acc = itemList[2]
            elif '+' in itemList[-1]:
                name = itemList[0]
                ratting = itemList[1]
                acc = itemList[2]
            else:
                name = itemList[0]
                ratting = itemList[1]
                note = itemList[2]
        elif len(itemList) == 4:
            name = itemList[0]
            ratting = itemList[1]
            acc = itemList[2]
            note = itemList[3]
        return name, ratting, acc, note

# Excel 제품명 정리 저장
    def setItem(self, path):
        itemCol = 2
        nameCol = 3
        rattingCol = 4
        accCol = 5
        noteCol = 6
        excelWB = op.load_workbook(path)
        excelWS = excelWB.active
        for _ in range(4):  excelWS.insert_cols(itemCol + 1)
        excelWS.cell(row=1, column=nameCol).value = 'name'
        excelWS.cell(row=1, column=rattingCol).value = 'ratting'
        excelWS.cell(row=1, column=accCol).value = 'acc'
        excelWS.cell(row=1, column=noteCol).value = 'note'
        maxRow = excelWS.max_row
        maxColumn = excelWS.max_column
        for i in range(2, maxRow + 1):
            name, ratting, acc, note = self.itemSepa(excelWS.cell(row=i, column=itemCol).value)
            excelWS.cell(row=i, column=nameCol).value = name
            excelWS.cell(row=i, column=rattingCol).value = ratting
            excelWS.cell(row=i, column=accCol).value = acc
            excelWS.cell(row=i, column=noteCol).value = note
        excelWB.save(path)
        excelWB.close()

# Excel 인쇄수량 데이터 변환 저장
    def setPrint(self, path):
        dataCols = [3, 4, 13, 30]
        startRow = 3
        quantityColumn = 3
        packagingColumn = 4
        printQuantity = 5

        excelAddress = path
        newExcel = excelAddress[0:-5] + 'NiceLabel용.xlsx'
        wb = op.load_workbook(excelAddress)
        ws = wb.active
        maxRow = ws.max_row
        maxColumn = ws.max_column

        for i in range(maxColumn, 0, -1):
            if not i in dataCols:
                ws.delete_cols(i)

        ws.cell(column=printQuantity, row=1).value = '인쇄수량'
        maxColumn = ws.max_column
        for i in range(maxRow, (startRow - 1), -1):
            if (ws.cell(row=i, column=quantityColumn).value % ws.cell(row=i, column=packagingColumn).value) != 0:
                if (ws.cell(row=i, column=quantityColumn).value // ws.cell(row=i, column=packagingColumn).value) == 0:
                    ws.cell(row=i, column=packagingColumn).value = ws.cell(row=i, column=quantityColumn).value
                    ws.cell(row=i, column=printQuantity).value = 1
                else:
                    ws.insert_rows(i + 1)
                    for j in range(1, maxColumn):
                        ws.cell(row=(i + 1), column=j).value = ws.cell(row=i, column=j).value
                    ws.cell(row=i, column=printQuantity).value = ws.cell(row=i, column=quantityColumn).value // ws.cell(
                        row=i, column=packagingColumn).value
                    ws.cell(row=i + 1, column=packagingColumn).value = ws.cell(row=i,
                                                                               column=quantityColumn).value % ws.cell(
                        row=i, column=packagingColumn).value
                    ws.cell(row=i + 1, column=printQuantity).value = 1
            else:
                ws.cell(row=i, column=printQuantity).value = ws.cell(row=i, column=quantityColumn).value // ws.cell(
                    row=i, column=packagingColumn).value
        ws.delete_rows(2)
        wb.save(newExcel)
        wb.close()
        return newExcel

# xls -> xlsx 변환
    def transXlsx(self, path):
        excelPath = path
        if path[-3:] == 'xls':
            excelPath = path[:-3] + 'xlsx'
            excelDF = pd.read_excel(path)
            excelDF.to_excel(excelPath, index=False)
        if os.path.exists(path):
            os.remove(path)
        return excelPath


class VersionInfo(QDialog):
    def __init__(self):
        self.logger = makeLogger()

        super().__init__()

        self.versionUI = uic.loadUi('Version.ui', self)
        self.setWindowTitle('version')
        self.setFixedSize(260, 144)


class ColorSetting(QDialog):
    def __init__(self):
        super().__init__()

        self.logger = makeLogger()
        self.initialCombobox = '색 그룹 추가'
        self.colorNames = []
        self.colorTxt = ''
        self.initUI()

    def initUI(self):
        try:
            with open('ColorName.txt', 'r') as f:
                self.colorNames = f.read().split()
            if self.colorNames:
                self.colorTxt = self.colorNames[0] + '.txt'
            else:
                self.colorNames.append(self.initialCombobox)
        except Exception as e:
            self.logger.critical(e)

        self.ui = uic.loadUi("ColorSetting.ui", self)
        self.setWindowTitle('ColorSetting')
        self.setFixedSize(540, 455)

        try:
            if self.initialCombobox not in self.colorNames:
                self.colorComboAction(self.colorNames[0])
            self.comboBoxMake(self.colorNames)
        except Exception as e:
            self.logger.critical(e)

        self.btn_itemAdd.clicked.connect(self.itemAdd)
        self.btn_itemRemove.clicked.connect(self.itemRemove)
        self.btn_colorItemSave.clicked.connect(self.colorItemSave)
        self.btn_colorNameAdd.clicked.connect(self.colorNameAdd)
        self.btn_colorNameRemove.clicked.connect(self.colorNameRemove)
        self.btn_itemAdd.setEnabled(False)
        self.btn_colorNameAdd.setEnabled(False)
        self.itemText.textChanged.connect(self.btn_itemAdd_setEnable)
        self.colorNameText.textChanged.connect(self.btn_colorNameAdd_setEnable)

        self.colorComboBox.activated[str].connect(self.colorComboAction)

# ColorSetting Window 시작
        self.show()

    def btn_colorNameAdd_setEnable(self):
        if not self.colorNameText.text():
            self.btn_colorNameAdd.setEnabled(False)
        else:
            self.btn_colorNameAdd.setEnabled(True)

    def btn_itemAdd_setEnable(self):
        if not self.itemText.text():
            self.btn_itemAdd.setEnabled(False)
        else:
            self.btn_itemAdd.setEnabled(True)

    def colorNameRemove(self):
        try:
            if self.initialCombobox not in self.colorNames:
                os.remove(self.colorComboBox.currentText() + '.txt')
                if os.path.exists(self.colorComboBox.currentText() + '.txt'):
                    os.remove(self.colorComboBox.currentText() + '.txt')
                del self.colorNames[self.colorNames.index(self.colorComboBox.currentText())]
                with open('colorName.txt', 'w') as colorNameTxt:
                    colorNameTxt.write(' '.join(self.colorNames))
                self.colorComboBox.removeItem(self.colorComboBox.currentIndex())
                if not self.colorNames:
                    self.colorNames.append(self.initialCombobox)
                    self.colorComboBox.addItem(self.initialCombobox)
                    self.itemList.clear()
                else:
                    self.colorComboAction(self.colorNames[0])
        except Exception as e:
            self.logger.critical(e)

    def colorNameAdd(self):
        if self.initialCombobox == self.colorComboBox.currentText():
            self.colorComboBox.removeItem(0)
            self.colorNames = []
        try:
            if self.colorNameText.text() not in self.colorNames:
                self.colorNames.append(self.colorNameText.text())
                self.colorTxt = self.colorNameText.text() + '.txt'
                with open('ColorName.txt', 'w') as saveColorNames:
                    saveColorNames.write(' '.join(self.colorNames))
                self.ColorTxtMake(self.colorNameText.text())
                self.colorComboBox.addItem(self.colorNameText.text())
            self.colorNameText.clear()
        except Exception as e:
            print(e)

    def ColorTxtMake(self, colorName):
        newColorTxt = colorName + '.txt'
        with open(newColorTxt, 'w') as f:
            pass

    def comboBoxMake(self, colorNames):
        for colorName in colorNames:
            self.colorComboBox.addItem(colorName)

# ListWdget 데이터 불러오기
    def listWdgetData(self):
        row = 0
        data = []
        while self.itemList.item(row):
            if self.itemList.item(row):
                data.append(self.itemList.item(row).text())
                row += 1
        return row, data

# ListWidget 데이터 저장하기
    def colorItemSave(self):
        try:
            if not self.initialCombobox in self.colorComboBox.currentText():
                row, data = self.listWdgetData()
                if row >= 1:
                    self.colorItemList = ''
                    with open(self.colorTxt, 'w') as colorTxtFile:
                        for item in data:
                            if not self.colorItemList:
                                self.colorItemList += item
                            else:
                                self.colorItemList = self.colorItemList + ' ' + item
                        colorTxtFile.write(self.colorItemList)
        except Exception as e:
            self.logger.critical(e)

# ComboBox에서 선택된 색의 데이터 ListWdget에 넣기
    def colorComboAction(self, text):
        if not self.initialCombobox == self.colorComboBox.currentText():
            self.itemList.clear()
            self.colorTxt = text + '.txt'
            with open(self.colorTxt, 'r') as colorTxtFile:
                dataList = colorTxtFile.read().split()
                for dataTxt in dataList:
                    self.itemList.addItem(dataTxt.strip())

# ListWdget에 데이터 넣기
    def itemAdd(self):
        if not self.initialCombobox == self.colorComboBox.currentText():
            try:
                self.itemtext = self.itemText.text()
                if self.itemtext.strip():
                    self.itemList.addItem(self.itemtext)
                    self.itemText.clear()
            except Exception as e:
                self.logger.critical(e)

# ListWdget에 데이터 지우기
    def itemRemove(self):
        if not self.initialCombobox == self.colorComboBox.currentText():
            removeItemRow = self.itemList.currentRow()
            self.itemList.takeItem(removeItemRow)

class MainWindow(QMainWindow):
    def __init__(self):
        self.logger = makeLogger()

        super().__init__()
        self.excelTrans = ExcelTrans()
        self.excelTrans.kaSetChangeDate = os.path.getmtime('KASET.xlsx')
        self.windowWeight = 600
        self.windowHight = 320
        self.setWindowTitle("NiceLabel Excel 변환")
        self.f = open('NiceLabelLocate.txt', 'r')
        self.appPath = self.f.read()
        self.excelFileFlag = ''
        self.setupUI()
        self.f.close()

    def setupUI(self):
        self.center()
        self.setFixedSize(self.windowWeight, self.windowHight)

        self.filelabel = QLabel('', self)
        self.filelabel.move(135, 30)
        self.filelabel.resize(500, 30)

        self.nclabel = QLabel(self.appPath, self)
        self.nclabel.move(135, 120)
        self.nclabel.resize(500, 30)

        self.excelTransMassage = QLabel('', self)
        self.excelTransMassage.move(135, 60)
        self.excelTransMassage.resize(500, 30)

        self.appOpenMassage = QLabel('', self)
        self.appOpenMassage.move(135, 150)
        self.appOpenMassage.resize(500, 30)

# 메뉴바 만들기
        ColorMenu = QAction('Color', self)
        ColorMenu.triggered.connect(self.setAction)

        version = QAction('Version', self)
        version.triggered.connect(self.versionAction)

        kaSet = QAction('KA set', self)
        kaSet.triggered.connect(self.goToKaSet)

        menubar = self.menuBar()

        Settingmenu = menubar.addMenu('&Setting')
        Settingmenu.addAction(ColorMenu)
        Settingmenu.addAction(kaSet)

        information = menubar.addMenu('&information')
        information.addAction(version)

# 메인윈도우 버튼
        btn_openfile = QPushButton("Excel File", self)
        btn_openfile.move(30, 30)
        btn_openfile.clicked.connect(self.btn_openFile)

        self.btn_exceltrans = QPushButton("변환", self)
        self.btn_exceltrans.move(30, 60)
        self.btn_exceltrans.clicked.connect(self.btn_excelTrans)
        self.btn_exceltrans.setEnabled(False)

        btn_exit = QPushButton("종료", self)
        btn_exit.move(30, 210)
        btn_exit.clicked.connect(QCoreApplication.instance().quit)

        btn_apppath = QPushButton("NiceLabal File", self)
        btn_apppath.move(30, 120)
        btn_apppath.clicked.connect(self.btn_appPath)

        self.btn_openapp = QPushButton("Open", self)
        self.btn_openapp.move(80, 150)
        self.btn_openapp.resize(50, 30)
        self.btn_openapp.clicked.connect(self.btn_openApp)

        self.btn_apppathsave = QPushButton("Save", self)
        self.btn_apppathsave.move(30, 150)
        self.btn_apppathsave.resize(50, 30)
        self.btn_apppathsave.clicked.connect(self.btn_appPathSave)

        if self.appPath:
            self.btn_apppathsave.setEnabled(True)
            self.btn_openapp.setEnabled(True)
        else:
            self.btn_apppathsave.setEnabled(False)
            self.btn_openapp.setEnabled(False)

        self.show()

# 메뉴바 이벤트
    def goToKaSet(self):
        try:
            self.excelTrans.kaDataPreSet()
            os.popen('KASET.xlsx')
        except Exception as e:
            print(e)

    def setAction(self):
        colorSetting = ColorSetting()
        colorSetting.exec_()

    def versionAction(self):
        versionInfo = VersionInfo()
        versionInfo.exec_()


# exe파일 위치 저장
    def btn_appPathSave(self):
        self.excelTrans.kaItemRowDictSet()
        with open('NiceLabelLocate.txt', 'w') as f:
            if self.appPath:
                f.truncate()
                f.write(self.appPath)
                self.appOpenMassage.setText('저장이 완료되었습니다.')
            else:
                self.appOpenMassage.setText('저장할 파일이 없습니다.')
                self.logger.info('저장할 파일이 없습니다.')

# exe파일 열기
    def btn_openApp(self):
        try:
            if self.appPath:
                os.popen(self.appPath)
                self.appOpenMassage.setText('파일이 열립니다.')
            else:
                self.appOpenMassage.setText('열 파일이 없습니다.')
                self.logger.info('열 파일이 없습니다.')
        except Exception as e:
            self.appOpenMassage.setText('파일이 이상합니다.')
            self.logger.critical('에러', e)

# exe 파일 위치 불러오기
    def btn_appPath(self):
        try:
            self.nicefile = QFileDialog.getOpenFileName(self, 'NiceLabel File', './')
            self.appPath = self.nicefile[0]
            self.nclabel.setText(self.appPath)
            if self.nclabel.text():
                self.btn_openapp.setEnabled(True)
                self.btn_apppathsave.setEnabled(True)
            else:
                self.btn_openapp.setEnabled(False)
                self.btn_apppathsave.setEnabled(False)
        except Exception as e:
            print('에러 발생 : ', e)

# Excel 파일 위치 불러오기
    def btn_openFile(self):
        self.openfile = QFileDialog.getOpenFileName(self, 'Open file', './')
        self.excelFileFlag = self.openfile[0]
        self.filelabel.setText(self.excelFileFlag)
        if 'xls' in self.filelabel.text()[-4:]:
            self.btn_exceltrans.setEnabled(True)
        else:
            self.btn_exceltrans.setEnabled(False)

# Excel 데이터 변환
    def btn_excelTrans(self):
        if self.excelFileFlag:
            excelPath = self.excelTrans.transXlsx(self.openfile[0])
            excelPath = self.excelTrans.setPrint(excelPath)
            self.excelTrans.setItem(excelPath)
            self.excelTrans.itemKaSave(excelPath)
            self.excelTransMassage.setText('변환이 완료되었습니다.')
        else:
            self.excelTransMassage.setText('파일이 없습니다.')

# 메인윈도우 중앙 배치
    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())


if __name__ == "__main__":
    app = QApplication(sys.argv)
    mainwindow = MainWindow()
    sys.exit(app.exec_())
